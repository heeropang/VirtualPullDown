#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import sys
import subprocess
import json
from openpyxl import Workbook
from requests import get
from time import sleep

###########################################
                                          #
filename        = 'ESMFold_structure'     # Name of your pdb file
                                          #
###########################################

input        = f"{filename}.pdb"
foldseek_url = "https://search.foldseek.com/api/ticket"

# Check if pdb exists
if not os.path.exists(input):
    print("Error: pdb not found.")
    sys.exit()

# Define the cURL command
curl_command = [
    "curl",
    "-X", "POST",
    "-F", f"q=@%s"%(input),
    "-F", "mode=3diaa",
    "-F", "database[]=afdb50",
   ## "-F", "database[]=afdb-swissprot",
   ## "-F", "database[]=afdb-proteome",
   ## "-F", "database[]=cath50",
   ## "-F", "database[]=mgnify_esm30",
   ## "-F", "database[]=pdb100",
   ## "-F", "database[]=gmgcl_id",
    "-F", "taxonomy[]=bacteria",  # Add your taxonomic filter here
    foldseek_url
]

# Execute the cURL command
try:
    output = subprocess.check_output(curl_command, universal_newlines=True)
    print(output)
except subprocess.CalledProcessError as e:
    print(f"Error executing cURL command: {e}")
    sys.exit()

# Parse the output to obtain the ticket ID
ticket = json.loads(output)
if not ticket:
    print("Error: Failed to retrieve the ticket ID.")
    sys.exit()

while True:
    if ticket['status'] == "ERROR":
        print("Job encountered an error.")
        sys.exit(0)
    elif ticket['status'] == "PENDING":
        print("Job is still running. Waiting...")
        sleep(10)
        results_url=f"https://search.foldseek.com/api/result/{ticket['id']}/0"
        results_response=get(results_url)
        try:
            results=results_response.json()
            break
        except json.JSONDecodeError:
            print("Error decoding JSON response:", results_response.text)
            continue
    elif ticket['status'] == "COMPLETE":
        print("Job completed... Preparing the list of hits...")
        results_url = f"https://search.foldseek.com/api/result/{ticket['id']}/0"
        results_response = get(results_url)
        try:
            results = results_response.json()
            break
        except json.JSONDecodeError:
            print("Error decoding JSON response:", results_response.text)
            sys.exit(1)
print(ticket['status'])
print("Job completed... Preparing the list of hits...")

wb        = Workbook()
ws        = wb.active
ws['A1']  = 'Target'
ws['B1']  = 'Description'
ws['C1']  = 'Probability'
ws['D1']  = 'E-value'
ws['E1']  = 'Scientific Name'
ws['F1']  = 'Sequence'

targets       = []
descriptions  = []
probabilities = []
evalues       = []
taxNames      = []
sequences     = []
for result in results['results']:
    db = result['db']
    alignments=result.get('alignments')
    if alignments:
        for alignment in alignments:
            target=alignment['target']
            prob = alignment['prob']
            eval = alignment['eval']
            taxName = alignment['taxName']
            tSeq = alignment['tSeq']
            targets.append(target.split('-F1')[0])
            descriptions.append(target.split('v4 ')[1])
            probabilities.append(prob)
            evalues.append(eval)
            taxNames.append(taxName)
            sequences.append(tSeq)

# Determine the maximum width for the description column
max_target_width = max(len(target) for target in targets)
max_description_width = max(len(description) for description in descriptions)

# Print the list of targets
for target, description, probability, ev, sciname, seq in zip(targets, descriptions, probabilities, evalues, taxNames, sequences):
    ws.append([target]+ [description]+[probability]+[ev] +[sciname] +[seq])
wb.save("%s_homologs.xlsx"%(filename))

subprocess.call("echo Please copy and paste the link below into a browser for the graphical interface...\n", shell=True)
subprocess.call("echo https://search.foldseek.com/result/%s/0"%(ticket['id']), shell=True)
