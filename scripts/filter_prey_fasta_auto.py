#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, sys
from os.path import join
import sys
from Bio import SeqIO
from openpyxl import Workbook
#############################################################
path            = "./"                                      # Directory of the integrase (bait) sequence
filename        = 'sequence'                                # Name of your fasta file
prey_size_limit = 400                                       # Provide size limit for prey sequences to avoid running out of memory in alphafold
bait_name       = sys.argv[1]                               #
with open(sys.argv[2], 'r') as f:                           #
    filter_start, filter_end= f.readline().strip().split()  #
filter_start    = int(filter_start)                         #
filter_end      = int(filter_end)                           #
#############################################################

f1_txt    = f"{filename}_check.txt" 
f2_txt    = f"{bait_name}.fasta"
txt_path  = os.path.join(path,f1_txt)
txt2_path = os.path.join(path,f2_txt)
record    = list(SeqIO.parse(txt_path, 'fasta'))
record2   = list(SeqIO.parse(txt2_path, 'fasta'))

wb        = Workbook()
ws        = wb.active
ws['A1']  = 'Locus_tag'
ws['B1']  = 'Gene_length'
ws['C1']  = 'Skip_%s'%(prey_size_limit)
ws['D1']  = 'start'
ws['E1']  = 'end'
ws['F1']  = 'selected'

for line in record:
    st=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[0]
    en=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[-1].split('>')[-1].split(')')[0].split(']')[0]
    if line.seq != record2[0].seq and len(line.seq) < prey_size_limit: 
        if int(st)<=filter_start and int(en) <= filter_start:
            pass
        elif int(st)<=filter_start and int(en) >= filter_start:
            of=open('%s.fa'%(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]),'w')
            of.write('>'+line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]+'\n')
            of.write(str(line.seq))
            of.close()
        elif int(st) >= filter_start and int(en) <= filter_end:
            of=open('%s.fa'%(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]),'w')
            of.write('>'+line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]+'\n')
            of.write(str(line.seq))
            of.close()
        elif int(st) >= filter_start and int(st) <= filter_end and int(en) >= filter_end:
            of=open('%s.fa'%(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]),'w')
            of.write('>'+line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]+'\n')
            of.write(str(line.seq))
            of.close()
        elif int(st) >= filter_end:
            pass
    elif line.seq != record2[0].seq and len(line.seq) > prey_size_limit:
        pass
for line in record:
    st=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[0]
    en=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[-1].split('>')[-1].split(')')[0].split(']')[0]
    if int(st)<= filter_start and int(en) <= filter_start:
        if len(line.seq) < prey_size_limit:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[len(line.seq)]\
                      +['no']\
                      +[st]\
                      +[en]\
                      +['no'])
        else:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[len(line.seq)]\
                      +['yes']\
                      +[st]\
                      +[en]\
                      +['no'])
    elif int(st) <= filter_start and int(en) >= filter_start:
        if len(line.seq) < prey_size_limit:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[len(line.seq)]\
                      +['no']\
                      +[st]\
                      +[en]\
                      +['yes'])
        else:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[len(line.seq)]\
                      +['yes']\
                      +[st]\
                      +[en]\
                      +['size_limit'])
    elif int(st) >= filter_start and int(en) <=filter_end:
        if len(line.seq) < prey_size_limit:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[len(line.seq)]\
                      +['no']\
                      +[st]\
                      +[en]\
                      +['yes'])
        else:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[len(line.seq)]\
                      +['yes']\
                      +[st]\
                      +[en]\
                      +['size_limit'])
    elif int(st) >= filter_start and int(st) <=filter_end and int(en) >=filter_end:
        if len(line.seq) < prey_size_limit:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[len(line.seq)]\
                      +['no']\
                      +[st]\
                      +[en]\
                      +['yes'])
        else:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[len(line.seq)]\
                      +['yes']\
                      +[st]\
                      +[en]\
                      +['size_limit'])
    elif int(line.description.split('..')[0].split("=")[-1].split('(')[-1].split('<')[-1]) >= filter_end:
        if len(line.seq) < prey_size_limit:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[len(line.seq)]\
                      +['no']\
                      +[st]\
                      +[en]\
                      +['no'])
        else:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[len(line.seq)]\
                      +['yes']\
                      +[st]\
                      +[en]\
                      +['no'])
wb.save("%s%s_filtered.xlsx"%(path,bait_name))
