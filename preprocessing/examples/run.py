#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, sys
from os.path import join
import sys
import re
import subprocess
import glob
from Bio import SeqIO
from openpyxl import Workbook

def combine_pairwise_batch(path, filenames, bait_name):
    filenames   = sorted(glob.glob("./fa/"+"*.fa"))
    f2_txt      = f"{bait_name}_bait_truncated.fasta"
    txt2_path   = os.path.join(path,f2_txt)    
    for fa in filenames:
        seq1    = list(SeqIO.parse(fa,'fasta'))
        seq2    = list(SeqIO.parse(txt2_path,'fasta'))
        seq_str = ''
        
        for seq2 in seq2:
            for seq1 in seq1:
                seq_str += '>' + seq1.id + '\n'
                seq_str += str(seq2.seq)+':'+str(seq1.seq) 
                of=open('%s_pair.fasta'%(fa),'w')
                of.write(seq_str)
                of.close()
                seq_str =''

def filter_prey_sequences(path, filename, prey_size_limit, bait_name, filter_start, filter_end):
    f1_txt      = f"{filename}_check.txt"
    f2_txt      = f"{bait_name}.fasta"
    txt_path    = os.path.join(path,f1_txt)
    txt2_path   = os.path.join(path,f2_txt)
    record      = list(SeqIO.parse(txt_path, 'fasta'))
    record2     = list(SeqIO.parse(txt2_path, 'fasta'))

    wb          = Workbook()
    ws          = wb.active
    ws['A1']    = 'Locus_tag'
    ws['B1']    = 'Gene_length'
    ws['C1']    = 'Skip_%s'%(prey_size_limit)
    ws['D1']    = 'start'
    ws['E1']    = 'end'
    ws['F1']    = 'selected'

    for line in record:
        st=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[0]
        en=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[-1].split('>')[-1].split(')')[0].split(']')[0]
        if line.seq != record2[0].seq and len(line.seq) < prey_size_limit: 
            if int(st)<=filter_start and int(en) <= filter_start:
                pass
            elif int(st)<=filter_start and int(en) >= filter_start:
                of=open('%s.fa'%(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]),'w')
                of.write('>'+line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]+'\n')
                of.write(str(line.seq))
                of.close()
            elif int(st) >= filter_start and int(en) <= filter_end:
                of=open('%s.fa'%(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]),'w')
                of.write('>'+line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]+'\n')
                of.write(str(line.seq))
                of.close()
            elif int(st) >= filter_start and int(st) <= filter_end and int(en) >= filter_end:
                of=open('%s.fa'%(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]),'w')
                of.write('>'+line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]+'\n')
                of.write(str(line.seq))
                of.close()
            elif int(st) >= filter_end:
                pass
        elif line.seq != record2[0].seq and len(line.seq) > prey_size_limit:
            pass
    for line in record:
        st=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[0]
        en=line.description.split('[location=')[-1].split('complement(')[-1].split('<')[-1].split('join(')[-1].split(',')[-1].split('..')[-1].split('>')[-1].split(')')[0].split(']')[0]
        if int(st)<= filter_start and int(en) <= filter_start:
            if len(line.seq) < prey_size_limit:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['no']\
                          +[st]\
                          +[en]\
                          +['no'])
            else:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['yes']\
                          +[st]\
                          +[en]\
                          +['no'])
        elif int(st) <= filter_start and int(en) >= filter_start:
            if len(line.seq) < prey_size_limit:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['no']\
                          +[st]\
                          +[en]\
                          +['yes'])
            else:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['yes']\
                          +[st]\
                          +[en]\
                          +['size_limit'])
        elif int(st) >= filter_start and int(en) <=filter_end:
            if len(line.seq) < prey_size_limit:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['no']\
                          +[st]\
                          +[en]\
                          +['yes'])
            else:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['yes']\
                          +[st]\
                          +[en]\
                          +['size_limit'])
        elif int(st) >= filter_start and int(st) <=filter_end and int(en) >=filter_end:
            if len(line.seq) < prey_size_limit:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['no']\
                          +[st]\
                          +[en]\
                          +['yes'])
            else:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['yes']\
                          +[st]\
                          +[en]\
                          +['size_limit'])
        elif int(line.description.split('..')[0].split("=")[-1].split('(')[-1].split('<')[-1]) >= filter_end:
            if len(line.seq) < prey_size_limit:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['no']\
                          +[st]\
                          +[en]\
                          +['no'])
            else:
                ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                          +[len(line.seq)]\
                          +['yes']\
                          +[st]\
                          +[en]\
                          +['no'])
    wb.save("%s%s_filtered.xlsx"%(path,bait_name))

def create_mastertable(path, filename, prey_size_limit, bait_name):
    f1_txt    = f"{filename}_check.txt"
    f2_txt    = f"{bait_name}.fasta"
    txt_path  = os.path.join(path,f1_txt)
    txt2_path = os.path.join(path,f2_txt)
    record    = list(SeqIO.parse(txt_path, 'fasta'))
    record2   = list(SeqIO.parse(txt2_path, 'fasta'))

    wb        = Workbook()
    ws        = wb.active
    ws['A1']  = 'Locus_tag'
    ws['B1']  = 'Gene_description'
    ws['C1']  = 'Gene_length'
    ws['D1']  = 'Skip_%s'%(prey_size_limit)
    ws['E1']  = 'start'
    ws['F1']  = 'end'
    ws['G1']  = '%s_info'%(bait_name)

    for line in record:
        if len(line.seq) <= prey_size_limit:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[line.description.split('protein=')[1].split(']')[0]]\
                      +[len(line.seq)]\
                      +['no']\
                      +[line.description.split('..')[0].split("=")[-1].split('(')[-1].split('<')[-1]]\
                      +[line.description.split('..')[1].split("]")[0].split(')')[0].split('>')[-1]])
        else:
            ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                      +[line.description.split('protein=')[1].split(']')[0]]\
                      +[len(line.seq)]\
                      +['yes']\
                      +[line.description.split('..')[0].split("=")[-1].split('(')[-1].split('<')[-1]]\
                      +[line.description.split('..')[1].split("]")[0].split(')')[0].split('>')[-1]])
    for line in record:
        if line.seq ==record2[0].seq:
            ws['H1']=str(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2])
    wb.save(f"{bait_name}_master.xlsx")

def identify_prophage_region(accession_number):
    url = f'http://phaster.ca/phaster_api?acc={accession_number}'
    result = subprocess.run(['wget', '-qO-', url], stdout=subprocess.PIPE)
    output = result.stdout.decode()
    keywords = re.findall(r'intact\(\d{1,10}\)|questionable\(\d{1,10}\)|incomplete\(\d{1,10}\)', output)
    ranges = re.findall(r'\d{1,10}-\d{1,10}', output)
    if not ranges:
        print("Prophage is not found")
    else:
        completeness = ' | '.join([f"{kw} {rg}" for kw, rg in zip(keywords, ranges)])
        print('Completeness and Position are :', completeness)
        filter_start = int(ranges[0].split('-')[0])
        filter_end = int(ranges[-1].split('-')[1])
        print('specified range from %s to %s will be applied unless manually specified'%(filter_start, filter_end))
        with open('range.txt', 'w') as f:
            f.write(str(filter_start) + ' ' + str(filter_end))
        return completeness, filter_start, filter_end
