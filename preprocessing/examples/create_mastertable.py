#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, sys
from os.path import join
import sys
from Bio import SeqIO
from openpyxl import Workbook
import subprocess

##################################
path            = "./"           # Directory of the integrase (bait) sequence
filename        = 'sequence'     # Name of your fasta file
prey_size_limit = 400            # Provide size limit for prey sequences to avoid running out of memory in alphafold
bait_name       = 'Bt24'         #
##################################

sed_cmd = "sed -e 's/\[db_xref=[^]]*\] //g' sequence.txt >sequence_check.txt"
subprocess.call(sed_cmd, shell=True)
        
f1_txt    = f"{filename}_check.txt" 
f2_txt    = f"{bait_name}.fasta"
txt_path  = os.path.join(path,f1_txt)
txt2_path = os.path.join(path,f2_txt)
record    = list(SeqIO.parse(txt_path, 'fasta'))
record2   = list(SeqIO.parse(txt2_path, 'fasta'))

wb        = Workbook()
ws        = wb.active
ws['A1']  = 'Locus_tag'
ws['B1']  = 'Gene_description'
ws['C1']  = 'Gene_length'
ws['D1']  = 'Skip_%s'%(prey_size_limit)
ws['E1']  = 'start'
ws['F1']  = 'end'
ws['G1']  = '%s_info'%(bait_name)

for line in record:
    if len(line.seq) <= prey_size_limit:
        ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                  +[line.description.split('protein=')[1].split(']')[0]]\
                  +[len(line.seq)]\
                  +['no']\
                  +[line.description.split('..')[0].split("=")[-1].split('(')[-1].split('<')[-1]]\
                  +[line.description.split('..')[1].split("]")[0].split(')')[0].split('>')[-1]])
    else:
        ws.append([line.description.split('[locus_tag=')[1].split('[protein')[0][:-2]]\
                  +[line.description.split('protein=')[1].split(']')[0]]\
                  +[len(line.seq)]\
                  +['yes']\
                  +[line.description.split('..')[0].split("=")[-1].split('(')[-1].split('<')[-1]]\
                  +[line.description.split('..')[1].split("]")[0].split(')')[0].split('>')[-1]])
for line in record:
    if line.seq ==record2[0].seq:
        ws['H1']=str(line.description.split('[locus_tag=')[1].split('[protein')[0][:-2])
        
wb.save("%s%s_master.xlsx"%(path,bait_name))
